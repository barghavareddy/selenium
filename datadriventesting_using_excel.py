import openpyxl
#path= "C:\Users\pc\Downloads\Financial Sample.xlsx"
path="C:\exeldata\Financial Sample.xlsx"
workbook=openpyxl.load_workbook(path)
sheet=workbook.active
r=sheet.max_row
c=sheet.max_column
for i in range(1,r+1):
    for j in range(1,c+1):
        print(sheet.cell(row=i,column=j).value,end="                         ")
    print()